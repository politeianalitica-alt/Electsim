#!/usr/bin/env python3
"""
Sprint G14 extra · build-china-coverage-baseline

Procesa el dataset People's Daily Coverage del repo China-Media-main (CFC v.1
· paper JCC 2025) para generar un baseline histórico ES-friendly de cobertura
china por país.

Para cada país calcula sobre los últimos 10 años (2013-2022):
  - mean_articles_year
  - std_articles_year (desviación estándar)
  - last_year_count + last_year
  - n_years_available

Esto habilita el detector de anomalías: si Xinhua hoy cubre un país muy por
encima o por debajo de su baseline 10-año, es señal de shift diplomático
chino que merece atención.

Input:  gits amigos/China-Media-main/Foreign Coverage by Country.xlsx
Output: apps/visual-oscar/data/china-pd-coverage-baseline.json

Idempotente · re-ejecutable cuando aparezca v.2 del dataset.
"""
import json
import sys
import statistics
from pathlib import Path
from openpyxl import load_workbook

# Repo principal (no el worktree) contiene `gits amigos/`
ROOT_CANDIDATES = [
    Path(__file__).resolve().parents[3],  # worktree root
    Path("/Users/antoniolegaz/Downloads/Politeria/electsim-espana 2"),  # repo principal
]
SRC = None
for r in ROOT_CANDIDATES:
    candidate = r / "gits amigos" / "China-Media-main" / "Foreign Coverage by Country.xlsx"
    if candidate.exists():
        SRC = candidate
        break
if SRC is None:
    SRC = ROOT_CANDIDATES[-1] / "gits amigos" / "China-Media-main" / "Foreign Coverage by Country.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "china-pd-coverage-baseline.json"

# Window de baseline · últimos 10 años de dataset
BASELINE_YEARS = list(range(2013, 2023))


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: dataset no encontrado en {SRC}", file=sys.stderr)
        return 1

    print(f"Input:  {SRC}")
    print(f"Output: {OUT}")

    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["PD Coverage"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # Encontrar indices columnas
    col_country = header.index("country")
    col_year = header.index("year")
    col_ccode = header.index("ccode")
    col_articles = header.index("article_count")

    # Agrupar por país
    by_country: dict[str, dict] = {}
    for row in rows[1:]:
        country = row[col_country]
        year = int(row[col_year]) if row[col_year] is not None else None
        ccode = row[col_ccode]
        articles = row[col_articles]
        if not country or year is None or articles is None:
            continue
        if not isinstance(articles, (int, float)):
            continue
        if country not in by_country:
            by_country[country] = {"iso3": ccode, "country": country, "years": {}}
        by_country[country]["years"][year] = float(articles)

    # Build baseline per country
    out = {}
    skipped = 0
    for country, data in by_country.items():
        years = data["years"]
        # Filtrar al window baseline
        window_values = [years[y] for y in BASELINE_YEARS if y in years and years[y] is not None]
        if len(window_values) < 3:
            skipped += 1
            continue  # insuficiente data para baseline confiable
        latest_year = max(years.keys())
        out[data["iso3"]] = {
            "country": country,
            "iso3": data["iso3"],
            "baseline_window": f"{BASELINE_YEARS[0]}-{BASELINE_YEARS[-1]}",
            "n_years_in_baseline": len(window_values),
            "mean_articles_year": round(statistics.mean(window_values), 2),
            "median_articles_year": round(statistics.median(window_values), 2),
            "std_articles_year": round(statistics.stdev(window_values), 2) if len(window_values) >= 2 else 0,
            "min_articles_year": min(window_values),
            "max_articles_year": max(window_values),
            "last_year": latest_year,
            "last_year_count": years.get(latest_year),
        }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "_meta": {
                    "source": "Politicization of People's Daily Coverage (CFC v.1) · paper JCC 2025",
                    "source_repo": "gits amigos/China-Media-main",
                    "baseline_window": f"{BASELINE_YEARS[0]}-{BASELINE_YEARS[-1]}",
                    "n_countries": len(out),
                    "n_skipped": skipped,
                    "methodology": "Per país: media + std + min/max de article_count en últimos 10 años del dataset. Habilita detección de anomalías: año actual ±2σ vs baseline = shift coverage chino.",
                    "limitations": "Dataset corta en 2022. Para detección actual usar como referencia histórica, no como dato vigente. NO mide tono editorial, sólo cantidad.",
                },
                "by_iso3": out,
            },
            f,
            ensure_ascii=False,
            separators=(",", ":"),
        )

    size_kb = OUT.stat().st_size / 1024
    spain = out.get("ESP")
    print(f"\n✓ {len(out)} países con baseline · {skipped} skipped · {size_kb:.1f} KB")
    if spain:
        print(f"  España (referencia): mean={spain['mean_articles_year']} std={spain['std_articles_year']} last={spain['last_year_count']} ({spain['last_year']})")
    # Top 5 por mean coverage
    top = sorted(out.values(), key=lambda c: -c["mean_articles_year"])[:5]
    print(f"  Top 5 cobertura: {', '.join(f'{c['country']}={c['mean_articles_year']}' for c in top)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
